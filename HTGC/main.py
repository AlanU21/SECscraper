from bs4 import BeautifulSoup
from io import StringIO
from datetime import datetime
import requests
import pandas as pd
import numpy as np
import regex as re
import logging
import unicodedata
import traceback
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

logging.basicConfig(filename='app.log', filemode='w', format='%(name)s - %(levelname)s - %(message)s', level=logging.DEBUG)



def setup_writer():
    return pd.ExcelWriter('cleaned_soi_tables.xlsx', engine='openpyxl')

def download_file(url):
    headers = {'User-Agent': "alanuthuppan@email.com"}
    try:
        response = requests.get(url, headers=headers)
        response.raise_for_status()
        return response.content
    except requests.RequestException as e:
        logging.error(f"Error downloading {url}: {str(e)}")
        return None

def parse_html(content):
    if content:
        return BeautifulSoup(content, 'html.parser')
    return None

def clean_html(soup):
    if soup:
        for tag in soup.recursiveChildGenerator():
            try:
                tag.attrs = {}
            except AttributeError:
                pass
        for linebreak in soup.find_all('br'):
            linebreak.extract()
    return soup

def get_qtr_date(date_format):
    month = int(date_format.strftime("%m"))
    year = date_format.strftime("%Y")
    quarter_ends = {12: "December 31", 9: "September 30", 6: "June 30", 3: "March 31"}
    return f"{quarter_ends.get(month, 'March 31')}, {year}"


###  EXTRACT  ###


def extract_tables(soup, qtr_date):
    tables = []
    count = 0
    logging.info(f"Current file: {qtr_date}")

    qtr_date = unicodedata.normalize('NFKD', qtr_date).replace('\xa0', ' ')
    phrase = 'CONSOLIDATED SCHEDU'
    date_pattern = re.compile(r'\b' + re.escape(qtr_date) + r'\b', re.IGNORECASE)
    all_dates_pattern = re.compile(r'\b(January|February|March|April|May|June|July|August|September|October|November|December)\s+\d{1,2},\s+\d{4}\b', re.IGNORECASE)

    for tag in soup.find_all(string=re.compile(phrase)):
        sibling = tag.parent.parent.find_next_sibling()
        if not sibling:
            sibling = tag.parent.find_next_sibling()
        if not sibling:
            sibling = tag.parent.parent.parent.find_next_sibling()

        if sibling and date_pattern.search(unicodedata.normalize('NFKD', sibling.get_text()).replace('\xa0', ' ')) and (phrase in tag.strip() and "ADVANCE" not in tag.strip()):
            logging.info(f"Found date matching {qtr_date} near tag: {tag}")
            html_table = sibling.find_next('table')
            if not html_table:
                html_table = sibling.find_next('div').find('table')

            # Additional check for incorrect date appearance
            if html_table:
                count += 1
                try:
                    new_table = pd.read_html(StringIO(str(html_table.prettify())), keep_default_na=False, skiprows=0, flavor='bs4')[0]
                    new_table = new_table.dropna(how='all', axis=0)
                    tables.append(new_table)

                except Exception as e:
                    logging.error(f"Failed to read HTML table near '{qtr_date}': {e}")
                
                try:
                    split_table = split_table_check(tables, html_table)
                    if split_table:
                        new_table = pd.read_html(StringIO(str(split_table.prettify())), keep_default_na=False, skiprows=0, flavor='bs4')[0]
                        new_table = new_table.dropna(how='all', axis=0)
                        tables.append(new_table)
                
                except Exception as e:
                    logging.error(f"Failed to read split table")

    if tables:
        for table in tables:
            table = table.map(lambda x: unicodedata.normalize('NFKD', x.strip().strip(u'\u200b').replace('—', '0').replace('%', '').replace('(', '').replace(')', '')) if isinstance(x, str) else x)
            table = table.replace(r'^\s*$', np.nan, regex=True).replace(r'^\s*\$\s*$', np.nan, regex=True).replace(r'^\s*\)\s*$', np.nan, regex=True)

    logging.info(f"# of tables extracted: {count}")
    return tables


def split_table_check(tables, html_table):
    if not tables or not html_table:
        return None

    last_table = tables[-1]
    if last_table.empty:
        return None

    last_row_first_cell = last_table.iloc[-1, 0]
    if isinstance(last_row_first_cell, str) and "Total: Debt Investments" in last_row_first_cell:
        next_table = html_table.find_next('table')
        return next_table


###  EXTRACT  ###





###  CLEAN  ###

def remove_nans(row):
    return pd.Series([x for x in row if pd.notna(x)])

def detect_headers(df):
    header_row = None
    headers = None
    for index, row in df.iterrows():
        if any("Portfolio" in str(value) for value in row):
            headers = row.values
            header_row = index
            return headers, header_row
        
    return headers, header_row


def detect_table_type(headers):
    debt_keywords = {
        re.compile(r'Maturity\s*Date', re.IGNORECASE),
        re.compile(r'Interest\s*Rate\s*and\s*Floor', re.IGNORECASE),
        re.compile(r'Principal\s*Amount', re.IGNORECASE)
    }
    equity_keywords = {
        re.compile(r'Acquisition\s*Date', re.IGNORECASE),
        re.compile(r'Series', re.IGNORECASE),
        re.compile(r'Shares', re.IGNORECASE)
    }
    
    if headers is not None:
        for header in headers:
            if isinstance(header, str):
                for pattern in debt_keywords:
                    if pattern.search(header):
                        logging.info("Detected DEBT table")
                        return 'Debt'
                for pattern in equity_keywords:
                    if pattern.search(header):
                        logging.info("Detected EQUITY table")
                        return 'Equity'
    
    logging.info("Detected UKNOWN table")
    return 'Unknown'






def clean_table(df):
    logging.info("Cleaning table")
    df = df.astype(str)
    df = df.map(lambda x: unicodedata.normalize('NFKD', str(x)).strip() if isinstance(x, str) else x)
    df = df[~df.map(lambda x: 'Subtotal' in x if isinstance(x, str) else False).any(axis=1)]
    df.replace(r'^\s*$', np.nan, regex=True, inplace=True)
    df = df.apply(lambda row: remove_nans(row), axis=1)    

    return df

def separate_frames(frames):
    cleaned_frames_debt = []
    cleaned_frames_equity = []
    prev_headers = pd.Series()
    for df in frames:
        try:
            cleaned_df = clean_table(df)
            headers, header_idx = detect_headers(cleaned_df)
            table_type = detect_table_type(headers)
            if table_type == 'Debt':
                cleaned_frames_debt.append(cleaned_df)
                prev_headers = headers
            elif table_type == 'Equity':
                cleaned_frames_equity.append(cleaned_df)
                prev_headers = headers
            elif table_type == 'Unknown':
                if len(prev_headers) == len(df.columns):
                    cleaned_frames_debt.append(cleaned_df)
            
        except Exception as e:
            logging.error(f"Error cleaning and separating frame: \n {df} \n {traceback.format_exc()}")
    
    logging.info(f"Number of cleaned debt frames: {len(cleaned_frames_debt)}")
    logging.info(f"Number of cleaned equity frames: {len(cleaned_frames_equity)}")

    return cleaned_frames_debt, cleaned_frames_equity









def consolidate_data(dataframes):
    if dataframes:
        master_table = dataframes[0]
        master_table_idx = detect_headers(master_table)[1]
        if master_table_idx is None:
            master_table_idx = 0
        master_table = master_table.iloc[master_table_idx:]

        for df in dataframes[1:]:
            header_idx = detect_headers(df)[1]
            master_table = pd.concat([master_table, df.iloc[header_idx+1:]], ignore_index=True)
        
        master_table.dropna(how='all', inplace=True)
        master_table.columns = master_table.iloc[0]
        master_table = master_table.iloc[1:].reset_index(drop=True)

        return master_table
    else:
        return pd.DataFrame



def master_cleaning(df):
    if not df.empty:

        #Currency concatenation
        currency_pattern = re.compile(r'[$€£¥₹]')
        for i, row in df.iterrows():
            j = 0
            while j < len(row):
                cell = row.iloc[j]
                if currency_pattern.match(str(cell)):
                    if j + 1 < len(row):
                        next_cell = row.iloc[j + 1]
                        df.iat[i, j] = str(cell) + str(next_cell)
                        df.iloc[i, j + 1:] = df.iloc[i, j + 2:].values.tolist() + [None]
                        row = df.iloc[i]
                j += 1
        
        #Getting rid of empty columns
        
        for i, col in enumerate(df.columns):
            if pd.isna(col):
                df = df.iloc[:, :i]
                break

        
        #Cell alignment
        
        rows_to_drop = []
        
        for i, row in df.iterrows():
            try:
                if isinstance(row.iloc[0], str) and "Total" in row.iloc[0]:
                    non_empty_values = row.dropna().iloc[1:].tolist()
                    num_values = len(non_empty_values)
                    if 'Footnotes' in df.columns:
                        start_index = num_values * -1 - 1
                        row.iloc[start_index:-1] = non_empty_values
                        row.iloc[1:len(non_empty_values)+1] = [np.nan] * len(non_empty_values)
                    else:
                        start_index = num_values * -1
                        row.iloc[start_index:] = non_empty_values
                        row.iloc[1:len(non_empty_values)+1] = [np.nan] * len(non_empty_values)
                
            
            except Exception as e:
                logging.error(f"'Total' row alignment error: {traceback.format_exc()}")
                logging.error(f"Start: {start_index}, Length of Non-Empty Values: {num_values}, Values: {non_empty_values}")
                logging.error(df)
                exit()
        
            try:
                if pd.notna(row.iloc[0]) and row.iloc[1:].isna().all():
                    rows_to_drop.append(i)

            except Exception as e:
                logging.error(f"Label rows removal error: {traceback.format_exc()}")
                logging.error(row)


        df = df.drop(rows_to_drop)

        #Missing Portfolio Company shift
        for i, row in df.iterrows():
            try:
                if row.iloc[0] in df.iloc[:, 1].values:
                    row.iloc[1:] = row.iloc[:-1]
                    row.iloc[0] = None
            
            except Exception as e:
                logging.error(f"Missing company name shift error: {traceback.format_exc()}")
                logging.error(row)

        

        return df
    


def extra_table_remover(df):
    if df is not None:
        for i, cell in enumerate(df.iloc[:, 0]):
            if isinstance(cell, str) and "Total: Debt Investments" in cell:
                df = df.iloc[:i+1]
                break
        return df


### CLEAN ###




### WRITE ###

def save_data(writer, data, sheet_name):
    if data is not None:
        try:
            data.to_excel(writer, sheet_name=sheet_name, index=False)
        except Exception as e:
            logging.error(f"Failed to save data for {sheet_name}: {traceback.format_exc()}")


def post_process_excel(file_path):
    try:
        workbook = load_workbook(file_path)
        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            for column in worksheet.columns:
                max_length = 0
                column = list(column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                worksheet.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width

            worksheet.freeze_panes = worksheet['A2']

        workbook.save(file_path)
        workbook.close()
        logging.info("Post-processing completed successfully.")
    except Exception as e:
        logging.error(f"Error during post-processing: {traceback.format_exc()}")



### WRITE ###

def scrape_data():
    writer = setup_writer()
    links = pd.read_excel("all_filings.xlsx")
    urls = links['Filings URL'].str.strip()
    date_reported = links['Reporting date']

    for i in range(len(urls)):
        content = download_file(urls[i])

        date_format = datetime.strptime(str(date_reported[i]), "%Y-%m-%d %H:%M:%S")
        qtr_date = get_qtr_date(date_format)

        if content:
            logging.info(f"Processing file: {urls[i]}")
            soup = parse_html(content)
            soup = clean_html(soup)
            tables = extract_tables(soup, qtr_date)
            cleaned_debt, cleaned_equity = separate_frames(tables)
            
            debt_sheet_name = f"{qtr_date}_DEBT"
            equity_sheet_name = f"{qtr_date}_EQUITY"
            
            consolidated_debt = consolidate_data(cleaned_debt)
            consolidated_equity = consolidate_data(cleaned_equity)

            consolidated_debt = master_cleaning(consolidated_debt)
            consolidated_debt = extra_table_remover(consolidated_debt)
            consolidated_equity = master_cleaning(consolidated_equity)

            
            save_data(writer, consolidated_debt, debt_sheet_name)
            save_data(writer, consolidated_equity, equity_sheet_name)
            
            logging.info(f"Data processing for {qtr_date} completed successfully.")
        else:
            logging.warning(f"No data was downloaded for URL {urls[i]}")

    writer._save()
    writer.close()

    post_process_excel('cleaned_soi_tables.xlsx')

    logging.info("All data has been processed and saved.")

def main():
    try:
        scrape_data()
    except Exception as e:
        logging.error(f"Unexpected error occurred: {traceback.format_exc()}")

if __name__ == "__main__":
    main()

