from bs4 import BeautifulSoup
from io import StringIO
from datetime import datetime
import requests
import pandas as pd
import numpy as np
import regex as re
import logging
import unicodedata
import traceback
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

logging.basicConfig(filename='app.log', filemode='w', format='%(name)s - %(levelname)s - %(message)s', level=logging.DEBUG)



def setup_writer():
    return pd.ExcelWriter('MVC_CLEANED_THRU_Q1_2014.xlsx', engine='openpyxl')

def download_file(url):
    headers = {'User-Agent': "alanuthuppan@email.com"}
    try:
        response = requests.get(url, headers=headers)
        response.raise_for_status()
        return response.content
    except requests.RequestException as e:
        logging.error(f"Error downloading {url}: {str(e)}")
        return None

def parse_html(content):
    if content:
        return BeautifulSoup(content, 'html.parser')
    return None

def clean_html(soup):
    if soup:
        for tag in soup.recursiveChildGenerator():
            try:
                tag.attrs = {}
            except AttributeError:
                pass
        for linebreak in soup.find_all('br'):
            linebreak.extract()
    return soup

def get_qtr_date(date_format):
    month = int(date_format.strftime("%m"))
    year = date_format.strftime("%Y")
    quarter_ends = {10: "October 31", 7: "July 31", 4: "April 30", 1: "January 31"}
    return f"{quarter_ends.get(month, 'March 31')}, {year}"




###  EXTRACT  ###

def normalize_text(text):
    return unicodedata.normalize('NFKD', text).replace('\xa0', ' ').strip()

def normalize_html_element(element):
    return unicodedata.normalize('NFKD', ' '.join(element.stripped_strings)).replace('\xa0', ' ').strip()

def create_flexible_pattern(some_str):
    parts = some_str.split()
    flexible_pattern = r'\s*'.join(map(re.escape, parts))
    return flexible_pattern


def extract_tables(soup, qtr_date):
    raw_tables = []
    count = 0
    logging.info(f"Current file: {qtr_date}")

    qtr_date = normalize_text(qtr_date)
    phrase = normalize_text('Consolidated Schedule of Investments')

    flexible_date_pattern = create_flexible_pattern(qtr_date)
    flexible_phrase_pattern = create_flexible_pattern(phrase)

    date_pattern = re.compile(r'.*\b' + flexible_date_pattern + r'\b.*', re.IGNORECASE)

    for tag in soup.find_all(string=re.compile(flexible_phrase_pattern)):
        parent = tag.parent
        date_found = False
        inside = False

        combined_text = normalize_html_element(parent)

        if date_pattern.search(combined_text):
            date_found = True
            logging.info(f"FOUND DATE matching {qtr_date} in parent tag")

        if not date_found:
            sibling = parent.find_next_sibling()
            while sibling and sibling.name in ['font', 'span', 'p']:
                combined_text += ' ' + normalize_html_element(sibling)
                if date_pattern.search(combined_text):
                    date_found = True
                    logging.info(f"FOUND DATE matching {qtr_date} in parent's sibling tags")
                    break
                sibling = sibling.find_next_sibling()
        
        if not date_found:
            grandparent = parent.parent
            if grandparent.name in ['tr', 'td']:
                great = grandparent.parent
                if great.name == 'tr':
                    sibling = great.find_next_sibling()
                    sibling_text = normalize_html_element(sibling)
                    if date_pattern.search(sibling_text):
                        date_found = True
                        inside = True
                        logging.info(f"FOUND DATE matching {qtr_date} INSIDE table")

        if not date_found:
            sibling = parent.parent.find_next_sibling()
            while sibling and sibling.name not in ['p', 'span']:
                sibling = sibling.find_next_sibling()
            
            if sibling and date_pattern.search(normalize_html_element(sibling)):
                date_found = True
                logging.info(f"FOUND DATE matching {qtr_date} in grandparent's sibling tags")
        
        if not date_found:
            sibling = parent.parent.parent.find_next_sibling()
            while sibling and sibling.name == 'p':
                sibling_text = normalize_html_element(sibling)
                if date_pattern.search(sibling_text):
                    date_found = True
                    logging.info(f"FOUND DATE matching {qtr_date} in great grandparent's sibling tags near tag: {tag}")
                    break
                sibling = sibling.find_next_sibling()
        
        if date_found and not inside:
            html_table = tag.parent.find_next('table')
            if html_table:
                try:
                    new_table = pd.read_html(StringIO(str(html_table.prettify())), keep_default_na=False, skiprows=0, flavor='bs4')[0]
                    new_table = new_table.dropna(how='all', axis=0)
                    new_table = new_table.dropna(how='all', axis=1)
                    raw_tables.append(new_table)

                except Exception as e:
                    logging.error(f"Failed to read OUTSIDE HTML table near '{tag}': {e}")
        

        elif date_found and inside:
            html_table = tag.parent.find_previous('table')
            if html_table:
                try:
                    new_table = pd.read_html(StringIO(str(html_table.prettify())), keep_default_na=False, skiprows=4, flavor='bs4')[0]
                    new_table = new_table.dropna(how='all', axis=0)
                    new_table = new_table.dropna(how='all', axis=1)

                    raw_tables.append(new_table)

                except Exception as e:
                    logging.error(f"Failed to read INSIDE HTML table near '{tag}': {e}")

        else:
            logging.info(f"Incorrect or no date found near parent tag: {parent}")
    
    if len(raw_tables) != 0:
        cleaned_tables = []
        for table in raw_tables:
            table = table.astype(str)
            table = table.replace(r'^\s*$', np.nan, regex=True)
            table = table.apply(lambda row: remove_nans(row), axis=1)

            currency_pattern = re.compile(r'[$€£¥₹]')
            percentage_pattern = re.compile(r'%')
            for i, row in table.iterrows():
                j = 0
                while j < len(row):
                    cell = row.iloc[j]
                    if currency_pattern.match(str(cell)):
                        if j + 1 < len(row):
                            next_cell = row.iloc[j + 1]
                            table.iat[i, j] = str(cell) + str(next_cell)
                            table.iloc[i, j + 1:] = table.iloc[i, j + 2:].values.tolist() + [None]
                            row = table.iloc[i]
                    
                    if percentage_pattern.match(str(cell)) and i not in [0, 1, 2]:
                        if j - 1 >= 0:
                            prev_cell = row.iloc[j - 1]
                            table.iat[i, j - 1] = str(prev_cell) + str(cell)
                            table.iloc[i, j:] = table.iloc[i, j + 1:].values.tolist() + [None]
                            row = table.iloc[i]

                    j += 1
                

            cleaned_tables.append(table)
            count += 1

        logging.info(f"# of tables extracted and initially cleaned: {count}")

        return cleaned_tables
    
    else:
        logging.error("NO TABLES FOUND for this filing\n")
        return None
    

###  EXTRACT  ###





###  CLEAN  ###

def remove_nans(row):
    return pd.Series([x for x in row if pd.notna(x)])

def detect_headers(df):
    header_row = None
    headers = None
    pattern = re.compile(r".*\bCOMPANY\b.*", re.IGNORECASE)

    for index, row in df.iterrows():
        if any(pattern.search(str(value)) for value in row):
            headers = row.values
            header_row = index
            return headers, header_row
        
    return None, None


def consolidate_data(dataframes, verify):
    if not dataframes:
        return pd.DataFrame()
    
    for i, df in enumerate(dataframes):
        empty_cols = df.columns[df.isna().all()]
        dataframes[i] = df.drop(empty_cols, axis=1)

    for i in reversed(range(len(dataframes))):
        df = dataframes[i]
        headers, header_row = detect_headers(df)
        if headers is None or header_row is None:
            logging.info("No headers found, removing irrelevant table")
            dataframes.pop(i)
    
    if not dataframes:
        return pd.DataFrame()

    master_table = dataframes[0]    
    headers, header_row = detect_headers(master_table)

    if len(dataframes) > 1:
        for df in dataframes[1:]:
            current_headers, current_header_row = detect_headers(df)
            if len(current_headers) != len(headers):
                logging.error(f"Headers do not match across tables. Skipping inconsistent table.\nCORRECT headers: {headers}\nCORRECT header length: {len(headers)}\nCurrent headers:{current_headers}\nCurrent header Length: {len(current_headers)}\n{df}\n")
                continue
            else:
                df = df.iloc[current_header_row+1:].reset_index(drop=True)                    
                df = df.reset_index(drop=True)
                master_table = pd.concat([master_table, df], ignore_index=True)
    
    master_table.dropna(how='all', inplace=True)
    master_table.columns = master_table.iloc[0]
    master_table = master_table.iloc[1:].reset_index(drop=True)

    logging.info(f"Consolidated {len(dataframes)} tables into single table with shape {master_table.shape}")

    if verify != 'nan':
        if len(dataframes) != verify:
            logging.error(f"MISMATCH - different # of tables extracted, correct # of tables should be: {verify}")
        else:
            logging.info("NUMBER OF TABLES VERIFIED")

    return master_table

def is_date(string):
    try:
        datetime.strptime(string, "%B %d, %Y")
        return True
    except ValueError:
        return False

def final_alignment(df, i):
    if i <= 16:
        df.columns = ['Company/Investment', 'Acquisition Date', 'Principal Amount', 'Cost', 'Fair Value', '% of Net Assets']
        for index, row in df.iterrows():
            if "Total" in str(row.iloc[0]):
                row.iloc[3:6] = row.iloc[1:4]
                row.iloc[1:3] = [np.nan, np.nan]

            second_cell = row.iloc[1]
            if not is_date(str(second_cell)) and "Total" not in str(row.iloc[0]):
                non_empty = row.iloc[1:].count()
                if non_empty == 3:
                    row.iloc[2:5] = row.iloc[1:4]
                    row.iloc[1] = np.nan
                elif non_empty == 2:
                    row.iloc[3:5] = row.iloc[1:3]
                    row.iloc[1:3] = [np.nan, np.nan]

        return df
    
    elif i <= 28:
        df.columns = ['Company/Investment', 'Principal Amount', 'Cost', 'Fair Value', '% of Net Assets']
        for index, row in df.iterrows():
            if "Total" in str(row.iloc[0]):
                row.iloc[2:5] = row.iloc[1:4]
                row.iloc[1] = np.nan
            
            else:
                non_empty = row.iloc[1:].count()
                if non_empty == 2:
                    row.iloc[2:4] = row.iloc[1:3]
                    row.iloc[1] = np.nan

        return df
    
    elif i <= 39:
        df.columns = ['Company/Investment', 'Industry', 'Principal Amount', 'Cost', 'Fair Value', '% of Net Assets']
        for index, row in df.iterrows():
            if "Total" in str(row.iloc[0]):
                row.iloc[3:6] = row.iloc[1:4]
                row.iloc[1:3] = [np.nan, np.nan]

            else:
                first_cell = row.iloc[0]
                second_cell = row.iloc[1]
                third_cell = row.iloc[2]

                if isinstance(first_cell, str) and str(first_cell) == "—":
                    row.iloc[3:5] = row.iloc[0:2]
                    row.iloc[0:2] = [np.nan, np.nan]

                elif isinstance(second_cell, str) and re.match(r'^\s*[$]?\d+(\.\d+)?\s*$', second_cell):
                    non_empty = row.iloc[1:].count()
                    if non_empty == 3:
                        row.iloc[2:5] = row.iloc[1:4]
                        row.iloc[1] = np.nan
                    elif non_empty == 2:
                        row.iloc[3:5] = row.iloc[1:3]
                        row.iloc[1:3] = [np.nan, np.nan]
                
                elif isinstance(third_cell, str) and str(third_cell) == "—":
                    row.iloc[3:5] = row.iloc[2:4]
                    row.iloc[2] = np.nan

        return df
    
    else:
        df.columns = ['Company', 'Industry', 'Investment', 'Principal Amount', 'Cost', 'Fair Value', '% of Net Assets']
        for index, row in df.iterrows():
            second_cell = row.iloc[1]
            if "Total" in str(row.iloc[0]) and second_cell != "—":
                row.iloc[4:7] = row.iloc[1:4]
                row.iloc[1:4] = [np.nan, np.nan, np.nan]

            elif "Total" in str(row.iloc[0]) and second_cell == "—":
                row.iloc[3:7] = row.iloc[1:5]
                row.iloc[1:3] = [np.nan, np.nan]
            
            else:
                if isinstance(second_cell, str) and re.match(r'^\s*[$]?\d+(\.\d+)?\s*$', second_cell):
                    row.iloc[2:6] = row.iloc[0:4]
                    row.iloc[0:2] = [np.nan, np.nan]
                    
        return df

def first_and_last_check(df):
    try:
        if df.empty or df.shape[0] < 2:
            return False

        first_row_pattern = re.compile(r"^\s*Senior\s+Secured\s+Notes\s*$", re.IGNORECASE)
        last_row_pattern = re.compile(r"^\s*Total\s+Investments\s*.*$", re.IGNORECASE)
        
        first_row_phrase = normalize_text(str(df.iloc[0, 0]))
        first_row_check = bool(first_row_pattern.match(first_row_phrase))

        last_row_phrase = normalize_text(str(df.iloc[-1, 0]))
        last_row_check = bool(last_row_pattern.search(last_row_phrase))

        logging.info(f"First cell: '{first_row_phrase}' - Match: {first_row_check}")
        logging.info(f"Last cell: '{last_row_phrase}' - Match: {last_row_check}")

        if first_row_check and last_row_check:
            logging.info(f"PASSED - TABLE VERIFIED\n")
            return True
        else:
            logging.info(f"FAILED - First cell: {first_row_phrase}\nLast cell: {last_row_phrase}\n")
            return False
    
    except Exception as e:
        logging.error(f"Error during first and last check: {traceback.format_exc()}")
        return False



### CLEAN ###




### WRITE ###

def save_data(writer, data, sheet_name):
    if data is not None:
        try:
            data.to_excel(writer, sheet_name=sheet_name, index=False)
        except Exception as e:
            logging.error(f"Failed to save data for {sheet_name}: {traceback.format_exc()}")


def post_process_excel(file_path):
    try:
        workbook = load_workbook(file_path)
        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            for row in worksheet.iter_rows():
                if row[0].value and isinstance(row[0].value, str) and "Total" in row[0].value:
                    for cell in row:
                        cell.font = Font(bold=True)

            worksheet.freeze_panes = worksheet['A2']

        workbook.save(file_path)
        workbook.close()
        logging.info("Post-processing completed successfully.")
    except Exception as e:
        logging.error(f"Error during post-processing: {traceback.format_exc()}")



### WRITE ###

def scrape_data():
    writer = setup_writer()
    links = pd.read_excel("all_filings.xlsx")
    urls = links['Filings URL'].str.strip()
    date_reported = links['Reporting date']
    verify = links['Correct # of tables']

    for i in range(len(urls)):
        if urls[i]:
            content = download_file(urls[i])
        else:
            continue

        try:
            date_format = datetime.strptime(str(date_reported[i]), "%Y-%m-%d %H:%M:%S")
            qtr_date = get_qtr_date(date_format)
        except Exception as e:
            logging.error(f"Error during date formatting: {traceback.format_exc()}")
            continue

        if content:
            logging.info(f"Processing file: {urls[i]}")
            soup = parse_html(content)
            soup = clean_html(soup)
            tables = extract_tables(soup, qtr_date)

            combined = consolidate_data(tables, verify[i])

            # combined = final_alignment(combined, i)

            # first_last = first_and_last_check(combined)
            # if not first_last:
            #     first_last = first_and_last_check(combined)

            #     if not first_last:
            #         logging.error(f"SOME TABLES MAY BE MISSING FOR QTR DATE: {qtr_date}\n")
            
            
            
            save_data(writer, combined, qtr_date)
            
            logging.info(f"Data processing for {qtr_date} completed successfully.\n")
        else:
            logging.warning(f"No data was downloaded for URL {urls[i]}")

    writer._save()
    writer.close()

    post_process_excel('MVC_CLEANED_THRU_Q1_2014.xlsx')

    logging.info("All data has been processed and saved.")

def main():
    try:
        scrape_data()
    except Exception as e:
        logging.error(f"Unexpected error occurred: {traceback.format_exc()}")

if __name__ == "__main__":
    main()
